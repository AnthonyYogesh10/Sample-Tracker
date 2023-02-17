import inspect
import logging

from openpyxl import Workbook ,load_workbook

class utils():
    def custom_logger(self,mode):
        #set class or method name from where it is called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(level=logging.DEBUG)
        # create console handler or file handler
        fh = logging.FileHandler(filename="automation.log",mode= mode)
        # add formatter to console or file handler
        formatter = logging.Formatter('%(asctime)s %(name)s %(levelname)s:%(message)s', datefmt='%d-%b-%y %H:%M:%S')
        # add formatter to console or file handler
        fh.setFormatter(formatter)
        # add console handller to logger
        logger.addHandler(fh)
        return logger


    def read_data_from_excel(self,filename, sheet):
        # empty list to store read data
        data_list = []
        wb = load_workbook(filename= filename)
        ws = wb[sheet]
        #find the lenth of the row and column
        row_count = ws.max_row
        col_count = ws.max_column

        for i in range(2, row_count + 1):
            row = []
            for j in range(1, col_count ):
                row.append(ws.cell(row=i, column=j).value)
            data_list.append(row)
        return data_list



# utlities are used to set common things which not related with driver
# base module are used to set common things which are related with driver
